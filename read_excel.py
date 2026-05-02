import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\Admin\Downloads\07.04.2026 - CONCREMAC_COMPARATIVO_SIKA.xlsx')
ws = wb.active

print(f"Sheet: {ws.title}")
print("\n--- K2:L11 (Dados do Traço) ---")
trace_data = {}
for i in range(2, 12):
    k_val = ws[f"K{i}"].value
    l_val = ws[f"L{i}"].value
    print(f"K{i}: {k_val} | L{i}: {l_val}")
    if k_val:
        trace_data[str(k_val)] = l_val

print("\n--- Colunas (primeiras 15) ---")
headers = [ws.cell(1, j).value for j in range(1, 16)]
print(headers)

print("\n--- Primeiras 3 linhas de dados ---")
for row_idx in range(2, 5):
    row_data = [ws.cell(row_idx, j).value for j in range(1, 16)]
    print(row_data)

print("\n--- Informações sobre fórmulas em K2:L11 ---")
for i in range(2, 12):
    cell_k = ws[f"K{i}"]
    cell_l = ws[f"L{i}"]
    if cell_k.data_type == 'f':
        print(f"K{i} é fórmula: {cell_k.value}")
    if cell_l.data_type == 'f':
        print(f"L{i} é fórmula: {cell_l.value}")
